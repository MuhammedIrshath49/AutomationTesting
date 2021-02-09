import openpyxl
from pageObjects.HomePage import Homepage
import time


class ExistingCard:
    jobOrder_xpath = "//body/div[@id='wrapper']/div[3]/div[1]/div[1]/div[1]/div[1]/input[1]"
    excelButton_xpath = "//input[@id='img']"
    calculateAmount_xpath = "//body/div[@id='wrapper']/div[3]/div[1]/div[1]/div[5]/div[1]/button[1]"
    buttonSubmit_xpath = "//body/div[@id='wrapper']/div[3]/div[1]/div[1]/div[5]/div[3]/button[1]"
    excpath = r"C:\Users\lukegoh\Desktop\Python Projects\SoftwareAutomationTesting\existingCard\Card Profile Top-up w topup valueTest 03_22.xlsx "
    excsheetName = 'data123'
    jobReqExistCard = "testec127"
    verifyStatus_xpath = "//body[1]/div[1]/div[3]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[3]/div[1]"
    rownum = 2

    def __init__(self, driver):
        self.driver = driver

    def createReqExistingCard(self, expected_price):
        self.driver.find_element_by_xpath(self.jobOrder_xpath).send_keys(self.jobReqExistCard)
        self.driver.find_element_by_xpath(self.excelButton_xpath).send_keys(self.excpath)
        self.driver.find_element_by_xpath(self.calculateAmount_xpath).click()
        subTotalAmt = self.driver.find_element_by_xpath(
            "//body[1]/div[1]/div[3]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[5]/div[1]").text
        totalCalAmt = self.driver.find_element_by_xpath(
            "//body[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[2]/label[1]").text.replace("Total amount: $", "")
        assert float(subTotalAmt) == float(totalCalAmt.split('$')[
                                               1]) == expected_price, "Price mismatch! Expected price is %s, subtotal is %s, total price is %s" % (
            expected_price, subTotalAmt, totalCalAmt)
        self.driver.find_element_by_xpath(self.buttonSubmit_xpath).click()
        time.sleep(8)

    def enterJobOrder(self):
        self.driver.find_element_by_xpath(self.jobOrder_xpath).send_keys(self.jobReqExistCard)

    def clickExcel(self):
        self.driver.find_element_by_xpath(self.excelButton_xpath).send_keys(self.excpath)

    def clickCalulateButton(self):
        self.driver.find_element_by_xpath(self.calculateAmount_xpath).click()

    def checkTopUpInfo(self):
        expected_price = 6.00
        subTotalAmt = self.driver.find_element_by_xpath(
            "//body[1]/div[1]/div[3]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").text
        totalCalAmt = self.driver.find_element_by_xpath("//label[contains(text(),'Total Amount: $6.00')]").text.replace(
            'Total amount: $', '')
        assert float(subTotalAmt) == float(totalCalAmt.split('$')[
                                               1]) == expected_price, "Price mismatch! Expected price is %s, subtotal " \
                                                                      "is %s, total price is %s" % (
                                                                          expected_price, subTotalAmt, totalCalAmt)

    def submitRequest(self):
        self.driver.find_element_by_xpath(self.buttonSubmit_xpath).click()
        time.sleep(5)
        expected_result = "Successfully submitted"
        result = self.driver.find_element_by_xpath("//body[1]/div[1]/div[3]/div[1]/div[1]/div[5]/label[1]").text
        assert result == expected_result, "Incorrect result. Expected result should be %s" % expected_result

    def checkStatus(self):
        hp = Homepage(self.driver)
        expected_status = "Pending"
        hp.clickviewRequest()
        self.driver.find_element_by_xpath("//input[@id='filterInput']").send_keys(self.jobReqNewCard)
        time.sleep(5)
        status = self.driver.find_element_by_xpath(self.verifyStatus_xpath).text
        assert status == expected_status, "Incorrect status, Expected status is %s" % expected_status

    def deleteRow(self):
        workbook = openpyxl.load_workbook(self.excpath)
        sheet = workbook.get_sheet_by_name(self.excsheetName)
        sheet.delete_rows(self.rownum, 1)
        workbook.save(self.excpath)

    def editValueInExcel(self):
        workbook = openpyxl.load_workbook(self.excpath)
        sheet = workbook.get_sheet_by_name(self.excsheetName)
        for r in range(2, 6):
            sheet.cell(row=r, col=2).value = "1"
        workbook.save(self.excpath)
